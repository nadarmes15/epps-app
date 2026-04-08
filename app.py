import os
import re
from datetime import datetime, date, timedelta
from functools import wraps
from io import BytesIO
from pathlib import Path
from tempfile import NamedTemporaryFile

from flask import (
    Flask,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine, text
from sqlalchemy.engine import Engine
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename
import xlrd

BASE_DIR = Path(__file__).resolve().parent
UPLOAD_DIR = BASE_DIR / "uploads"
LOCAL_DB_PATH = BASE_DIR / "instance" / "epps.db"

REQUIRED_COLUMNS = [
    "NRO. DNI",
    "NOMBRE AUXILIAR",
    "DESCRIPCION",
    "CANTIDAD",
    "FECHA DE MOVIMIENTO",
    "FECHA DE RENOVACION",
    "ESTADO",
]
OPTIONAL_COLUMNS = [
    "ESTADO MANUAL",
    "TIPO PLLA",
    "AREA",
    "CARGO",
    "OPERACIÓN",
    "OPERACION",
    "OPERACIÓN ",
]
ALLOWED_EXTENSIONS = {"xlsx", "xlsm", "xls"}


def get_database_url() -> str:
    raw = os.environ.get("DATABASE_URL", "").strip()
    if raw:
        if raw.startswith("postgres://"):
            raw = raw.replace("postgres://", "postgresql+psycopg://", 1)
        elif raw.startswith("postgresql://") and "+psycopg" not in raw:
            raw = raw.replace("postgresql://", "postgresql+psycopg://", 1)
        return raw
    (BASE_DIR / "instance").mkdir(exist_ok=True)
    return f"sqlite:///{LOCAL_DB_PATH}"


def create_db_engine() -> Engine:
    url = get_database_url()
    kwargs = {"future": True, "pool_pre_ping": True}
    if url.startswith("sqlite:///"):
        kwargs["connect_args"] = {"check_same_thread": False}
    return create_engine(url, **kwargs)


engine = create_db_engine()
app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "transportes-libertad-epps-2026")
app.config["UPLOAD_FOLDER"] = str(UPLOAD_DIR)
app.config["MAX_CONTENT_LENGTH"] = 25 * 1024 * 1024

UPLOAD_DIR.mkdir(exist_ok=True)


ADMIN_TABLE_SQL = """
CREATE TABLE IF NOT EXISTS admin_user (
    id INTEGER GENERATED ALWAYS AS IDENTITY PRIMARY KEY,
    username VARCHAR(100) NOT NULL UNIQUE,
    password_hash TEXT NOT NULL,
    created_at VARCHAR(30) NOT NULL
)
"""

EPP_TABLE_SQL = """
CREATE TABLE IF NOT EXISTS epp_records (
    id INTEGER GENERATED ALWAYS AS IDENTITY PRIMARY KEY,
    dni VARCHAR(20) NOT NULL,
    nombre_auxiliar TEXT,
    descripcion TEXT,
    cantidad TEXT,
    fecha_movimiento TEXT,
    fecha_renovacion TEXT,
    estado TEXT,
    estado_manual TEXT,
    tipo_plla TEXT,
    area TEXT,
    cargo TEXT,
    operacion TEXT,
    archivo_origen TEXT,
    fecha_importacion TEXT NOT NULL,
    renovacion_sort VARCHAR(20),
    movimiento_sort VARCHAR(20),
    estado_visual TEXT
)
"""


def db_execute(sql: str, params: dict | None = None):
    with engine.begin() as conn:
        return conn.execute(text(sql), params or {})


def db_fetchone(sql: str, params: dict | None = None):
    with engine.begin() as conn:
        result = conn.execute(text(sql), params or {})
        row = result.mappings().first()
        return dict(row) if row else None


def db_fetchall(sql: str, params: dict | None = None):
    with engine.begin() as conn:
        result = conn.execute(text(sql), params or {})
        return [dict(row) for row in result.mappings().all()]


def init_db():
    with engine.begin() as conn:
        dialect = engine.dialect.name
        if dialect == "sqlite":
            conn.execute(
                text(
                    """
                    CREATE TABLE IF NOT EXISTS admin_user (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        username TEXT NOT NULL UNIQUE,
                        password_hash TEXT NOT NULL,
                        created_at TEXT NOT NULL
                    )
                    """
                )
            )
            conn.execute(
                text(
                    """
                    CREATE TABLE IF NOT EXISTS epp_records (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        dni TEXT NOT NULL,
                        nombre_auxiliar TEXT,
                        descripcion TEXT,
                        cantidad TEXT,
                        fecha_movimiento TEXT,
                        fecha_renovacion TEXT,
                        estado TEXT,
                        estado_manual TEXT,
                        tipo_plla TEXT,
                        area TEXT,
                        cargo TEXT,
                        operacion TEXT,
                        archivo_origen TEXT,
                        fecha_importacion TEXT NOT NULL,
                        renovacion_sort TEXT,
                        movimiento_sort TEXT,
                        estado_visual TEXT
                    )
                    """
                )
            )
        else:
            conn.execute(text(ADMIN_TABLE_SQL))
            conn.execute(text(EPP_TABLE_SQL))

        username = os.environ.get("ADMIN_USERNAME", "admin")
        password = os.environ.get("ADMIN_PASSWORD", "Admin123*")
        exists = conn.execute(
            text("SELECT id FROM admin_user WHERE username = :username"),
            {"username": username},
        ).mappings().first()
        if not exists:
            conn.execute(
                text(
                    "INSERT INTO admin_user (username, password_hash, created_at) VALUES (:username, :password_hash, :created_at)"
                ),
                {
                    "username": username,
                    "password_hash": generate_password_hash(password),
                    "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                },
            )


def normalize_col(value: str) -> str:
    normalized = " ".join(str(value).replace("\n", " ").replace("\r", " ").strip().upper().split())
    aliases = {
        "FECHA MOVIMIENTO": "FECHA DE MOVIMIENTO",
        "OPERACION": "OPERACIÓN",
        "OPERACIÓN ": "OPERACIÓN",
    }
    return aliases.get(normalized, normalized)


def allowed_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def login_required(view):
    @wraps(view)
    def wrapped_view(*args, **kwargs):
        if not session.get("admin_logged_in"):
            flash("Debe iniciar sesión para acceder al panel administrador.", "warning")
            return redirect(url_for("admin_login"))
        return view(*args, **kwargs)

    return wrapped_view


def clean_text(value) -> str:
    if value is None:
        return ""
    text_value = str(value).strip()
    return "" if text_value.lower() == "none" else text_value


def parse_date_value(value, datemode=None):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    if isinstance(value, (int, float)):
        if datemode is not None:
            try:
                parts = xlrd.xldate_as_tuple(float(value), datemode)
                return datetime(*parts[:6])
            except Exception:
                pass
        try:
            base = datetime(1899, 12, 30)
            return base + timedelta(days=float(value))
        except Exception:
            return None

    text_value = str(value).strip()
    if not text_value or text_value.lower() in {"nan", "nat", "none"}:
        return None

    for fmt in (
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%Y-%m-%d",
        "%d/%m/%Y %H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%m/%d/%Y",
        "%d.%m.%Y",
    ):
        try:
            return datetime.strptime(text_value, fmt)
        except ValueError:
            continue

    try:
        return datetime.fromisoformat(text_value.replace("Z", ""))
    except Exception:
        return None


def to_display_date(value, datemode=None):
    parsed = parse_date_value(value, datemode)
    if parsed is None:
        return clean_text(value)
    return parsed.strftime("%d/%m/%Y")


def to_sort_date(value, datemode=None):
    parsed = parse_date_value(value, datemode)
    if parsed is None:
        return ""
    return parsed.strftime("%Y-%m-%d")


def _load_excel_rows(file_path: str):
    ext = file_path.rsplit(".", 1)[1].lower()
    if ext in {"xlsx", "xlsm"}:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        worksheet = workbook[workbook.sheetnames[0]]
        rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
        return rows, None
    if ext == "xls":
        book = xlrd.open_workbook(file_path)
        sheet = book.sheet_by_index(0)
        rows = [sheet.row_values(i) for i in range(sheet.nrows)]
        return rows, book.datemode
    raise ValueError("Formato no soportado.")


def detect_header_row_from_rows(rows):
    required_norm = {normalize_col(col) for col in REQUIRED_COLUMNS}
    for idx, row in enumerate(rows[:12]):
        row_values = {normalize_col(v) for v in row if clean_text(v)}
        if required_norm.issubset(row_values):
            return idx
    return 0


def read_excel_file(file_path: str):
    rows, datemode = _load_excel_rows(file_path)
    if not rows:
        raise ValueError("El archivo no contiene hojas o filas.")

    header_row = detect_header_row_from_rows(rows)
    headers = [normalize_col(v) for v in rows[header_row]]
    if not any(headers):
        raise ValueError("No se pudo detectar la fila de encabezados.")

    seen = set()
    normalized_headers = []
    for header in headers:
        if header and header not in seen:
            normalized_headers.append(header)
            seen.add(header)
        else:
            normalized_headers.append("")

    required_norm = [normalize_col(c) for c in REQUIRED_COLUMNS]
    missing = [col for col in required_norm if col not in normalized_headers]
    if missing:
        raise ValueError("Faltan columnas obligatorias: " + ", ".join(missing))

    keep_cols = list(
        dict.fromkeys(required_norm + [normalize_col(c) for c in OPTIONAL_COLUMNS if normalize_col(c) in normalized_headers])
    )

    data_rows = []
    for raw_row in rows[header_row + 1 :]:
        row_dict = {}
        for idx, header in enumerate(normalized_headers):
            if not header or header not in keep_cols:
                continue
            value = raw_row[idx] if idx < len(raw_row) else ""
            row_dict[header] = value

        if not any(clean_text(v) for v in row_dict.values()):
            continue

        dni_raw = clean_text(row_dict.get("NRO. DNI", ""))
        dni = re.sub(r"\D", "", dni_raw)[:8]
        if not dni:
            continue

        ren_sort = to_sort_date(row_dict.get("FECHA DE RENOVACION"), datemode)
        mov_sort = to_sort_date(row_dict.get("FECHA DE MOVIMIENTO"), datemode)
        ren_display = to_display_date(row_dict.get("FECHA DE RENOVACION"), datemode)
        mov_display = to_display_date(row_dict.get("FECHA DE MOVIMIENTO"), datemode)

        estado = clean_text(row_dict.get("ESTADO", ""))
        estado_manual = clean_text(row_dict.get("ESTADO MANUAL", ""))
        estado_visual = estado if estado else "SIN ESTADO"

        ren_dt = parse_date_value(row_dict.get("FECHA DE RENOVACION"), datemode)
        if ren_dt is not None:
            delta_days = (ren_dt.date() - date.today()).days
            if delta_days < 0:
                estado_visual = "VENCIDO"
            elif delta_days <= 30:
                estado_visual = "POR VENCER"
            elif estado_visual.upper() in {"", "SIN ESTADO"}:
                estado_visual = "VIGENTE"

        if estado_manual:
            estado_visual = estado_manual

        operacion = clean_text(row_dict.get("OPERACIÓN") or row_dict.get("OPERACION") or row_dict.get("OPERACIÓN ") or "")

        data_rows.append(
            {
                "NRO. DNI": dni,
                "NOMBRE AUXILIAR": clean_text(row_dict.get("NOMBRE AUXILIAR", "")),
                "DESCRIPCION": clean_text(row_dict.get("DESCRIPCION", "")),
                "CANTIDAD": clean_text(row_dict.get("CANTIDAD", "")),
                "FECHA DE MOVIMIENTO_DISPLAY": mov_display,
                "FECHA DE RENOVACION_DISPLAY": ren_display,
                "ESTADO": estado,
                "ESTADO MANUAL": estado_manual,
                "TIPO PLLA": clean_text(row_dict.get("TIPO PLLA", "")),
                "AREA": clean_text(row_dict.get("AREA", "")),
                "CARGO": clean_text(row_dict.get("CARGO", "")),
                "OPERACIÓN": operacion,
                "FECHA DE MOVIMIENTO_SORT": mov_sort,
                "FECHA DE RENOVACION_SORT": ren_sort,
                "ESTADO_VISUAL": estado_visual,
            }
        )

    if not data_rows:
        raise ValueError("No se encontraron registros válidos con DNI.")

    return data_rows


def replace_records_from_rows(rows, original_filename: str):
    now = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    with engine.begin() as conn:
        conn.execute(text("DELETE FROM epp_records"))
        stmt = text(
            """
            INSERT INTO epp_records (
                dni, nombre_auxiliar, descripcion, cantidad, fecha_movimiento,
                fecha_renovacion, estado, estado_manual, tipo_plla, area,
                cargo, operacion, archivo_origen, fecha_importacion,
                renovacion_sort, movimiento_sort, estado_visual
            ) VALUES (
                :dni, :nombre_auxiliar, :descripcion, :cantidad, :fecha_movimiento,
                :fecha_renovacion, :estado, :estado_manual, :tipo_plla, :area,
                :cargo, :operacion, :archivo_origen, :fecha_importacion,
                :renovacion_sort, :movimiento_sort, :estado_visual
            )
            """
        )
        for row in rows:
            conn.execute(
                stmt,
                {
                    "dni": row.get("NRO. DNI", ""),
                    "nombre_auxiliar": row.get("NOMBRE AUXILIAR", ""),
                    "descripcion": row.get("DESCRIPCION", ""),
                    "cantidad": row.get("CANTIDAD", ""),
                    "fecha_movimiento": row.get("FECHA DE MOVIMIENTO_DISPLAY", ""),
                    "fecha_renovacion": row.get("FECHA DE RENOVACION_DISPLAY", ""),
                    "estado": row.get("ESTADO", ""),
                    "estado_manual": row.get("ESTADO MANUAL", ""),
                    "tipo_plla": row.get("TIPO PLLA", ""),
                    "area": row.get("AREA", ""),
                    "cargo": row.get("CARGO", ""),
                    "operacion": row.get("OPERACIÓN", ""),
                    "archivo_origen": original_filename,
                    "fecha_importacion": now,
                    "renovacion_sort": row.get("FECHA DE RENOVACION_SORT", ""),
                    "movimiento_sort": row.get("FECHA DE MOVIMIENTO_SORT", ""),
                    "estado_visual": row.get("ESTADO_VISUAL", ""),
                },
            )


def get_dashboard_stats():
    base = db_fetchone(
        """
        SELECT COUNT(*) AS total_registros,
               COUNT(DISTINCT dni) AS total_trabajadores,
               COALESCE(MAX(fecha_importacion), 'Sin carga') AS ultima_importacion,
               COALESCE(MAX(archivo_origen), 'Sin archivo') AS archivo_origen
        FROM epp_records
        """
    ) or {}

    estados = db_fetchone(
        """
        SELECT
            SUM(CASE WHEN UPPER(estado_visual) = 'VENCIDO' THEN 1 ELSE 0 END) AS vencidos,
            SUM(CASE WHEN UPPER(estado_visual) = 'POR VENCER' THEN 1 ELSE 0 END) AS por_vencer,
            SUM(CASE WHEN UPPER(estado_visual) IN ('ENTREGADO', 'VIGENTE') THEN 1 ELSE 0 END) AS vigentes
        FROM epp_records
        """
    ) or {}

    return {
        "total_registros": base.get("total_registros", 0) or 0,
        "total_trabajadores": base.get("total_trabajadores", 0) or 0,
        "ultima_importacion": base.get("ultima_importacion", "Sin carga"),
        "archivo_origen": base.get("archivo_origen", "Sin archivo"),
        "vencidos": estados.get("vencidos", 0) or 0,
        "por_vencer": estados.get("por_vencer", 0) or 0,
        "vigentes": estados.get("vigentes", 0) or 0,
    }


@app.route("/", methods=["GET", "POST"])
def home():
    records = []
    dni = ""
    nombre = ""
    resumen = None

    if request.method == "POST":
        dni = "".join(ch for ch in request.form.get("dni", "") if ch.isdigit())
        if len(dni) != 8:
            flash("Ingrese un DNI válido de 8 dígitos.", "error")
        else:
            records = db_fetchall(
                """
                SELECT dni, nombre_auxiliar, descripcion, cantidad,
                       fecha_movimiento, fecha_renovacion, estado, estado_visual,
                       area, cargo, operacion
                FROM epp_records
                WHERE dni = :dni
                ORDER BY renovacion_sort ASC, descripcion ASC
                """,
                {"dni": dni},
            )
            if records:
                nombre = records[0]["nombre_auxiliar"]
                resumen = {
                    "total_items": len(records),
                    "vencidos": sum(1 for r in records if (r.get("estado_visual") or "").upper() == "VENCIDO"),
                    "por_vencer": sum(1 for r in records if (r.get("estado_visual") or "").upper() == "POR VENCER"),
                    "vigentes": sum(
                        1
                        for r in records
                        if (r.get("estado_visual") or "").upper() in {"ENTREGADO", "VIGENTE"}
                    ),
                }
            else:
                flash("No se encontraron registros para ese DNI.", "warning")

    return render_template("home.html", records=records, dni=dni, nombre=nombre, resumen=resumen)


@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        user = db_fetchone("SELECT * FROM admin_user WHERE username = :username", {"username": username})
        if user and check_password_hash(user["password_hash"], password):
            session.clear()
            session["admin_logged_in"] = True
            session["admin_username"] = username
            return redirect(url_for("admin_dashboard"))
        flash("Usuario o contraseña incorrectos.", "error")
    return render_template("admin_login.html")


@app.route("/admin/logout")
def admin_logout():
    session.clear()
    return redirect(url_for("admin_login"))


@app.route("/admin", methods=["GET", "POST"])
@login_required
def admin_dashboard():
    if request.method == "POST":
        excel = request.files.get("excel_file")
        if not excel or excel.filename == "":
            flash("Seleccione un archivo Excel.", "error")
            return redirect(url_for("admin_dashboard"))
        if not allowed_file(excel.filename):
            flash("Formato no permitido. Use .xlsx, .xlsm o .xls.", "error")
            return redirect(url_for("admin_dashboard"))

        filename = secure_filename(excel.filename)
        temp_suffix = Path(filename).suffix or ".xlsx"
        with NamedTemporaryFile(delete=False, suffix=temp_suffix) as tmp:
            excel.save(tmp.name)
            temp_path = tmp.name

        try:
            rows = read_excel_file(temp_path)
            final_name = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{filename}"
            replace_records_from_rows(rows, final_name)
            flash(f"Archivo cargado correctamente. Registros importados: {len(rows)}", "success")
        except Exception as exc:
            flash(f"No se pudo importar el archivo: {exc}", "error")
        finally:
            try:
                os.remove(temp_path)
            except OSError:
                pass
        return redirect(url_for("admin_dashboard"))

    filtro_dni = "".join(ch for ch in request.args.get("dni", "") if ch.isdigit())
    filtro_estado = request.args.get("estado", "").strip().upper()

    query = """
        SELECT dni, nombre_auxiliar, descripcion, fecha_renovacion, estado_visual, archivo_origen
        FROM epp_records
        WHERE 1=1
    """
    params = {}
    if filtro_dni:
        query += " AND dni = :dni"
        params["dni"] = filtro_dni
    if filtro_estado:
        query += " AND UPPER(estado_visual) = :estado_visual"
        params["estado_visual"] = filtro_estado
    query += " ORDER BY renovacion_sort ASC, id DESC LIMIT 25"

    recent_records = db_fetchall(query, params)
    stats = get_dashboard_stats()
    return render_template(
        "admin_dashboard.html",
        stats=stats,
        recent_records=recent_records,
        filtro_dni=filtro_dni,
        filtro_estado=filtro_estado,
    )


@app.route("/admin/change-password", methods=["POST"])
@login_required
def change_password():
    current_password = request.form.get("current_password", "")
    new_password = request.form.get("new_password", "")
    confirm_password = request.form.get("confirm_password", "")

    if len(new_password) < 8:
        flash("La nueva contraseña debe tener al menos 8 caracteres.", "error")
        return redirect(url_for("admin_dashboard"))
    if new_password != confirm_password:
        flash("La confirmación de contraseña no coincide.", "error")
        return redirect(url_for("admin_dashboard"))

    user = db_fetchone("SELECT * FROM admin_user WHERE username = :username", {"username": session["admin_username"]})
    if not user or not check_password_hash(user["password_hash"], current_password):
        flash("La contraseña actual es incorrecta.", "error")
        return redirect(url_for("admin_dashboard"))

    db_execute(
        "UPDATE admin_user SET password_hash = :password_hash WHERE id = :id",
        {"password_hash": generate_password_hash(new_password), "id": user["id"]},
    )
    flash("Contraseña actualizada correctamente.", "success")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/exportar")
@login_required
def admin_exportar():
    records = db_fetchall(
        """
        SELECT dni, nombre_auxiliar, descripcion, cantidad, fecha_movimiento,
               fecha_renovacion, estado, estado_manual, tipo_plla, area,
               cargo, operacion, estado_visual, archivo_origen, fecha_importacion
        FROM epp_records
        ORDER BY dni, renovacion_sort, descripcion
        """
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "EPPS"
    headers = [
        "NRO. DNI", "NOMBRE AUXILIAR", "DESCRIPCION", "CANTIDAD",
        "FECHA DE MOVIMIENTO", "FECHA DE RENOVACION", "ESTADO", "ESTADO MANUAL",
        "TIPO PLLA", "AREA", "CARGO", "OPERACIÓN", "ESTADO VISUAL",
        "ARCHIVO ORIGEN", "FECHA IMPORTACION",
    ]
    worksheet.append(headers)
    for row in records:
        worksheet.append([
            row.get("dni"), row.get("nombre_auxiliar"), row.get("descripcion"), row.get("cantidad"),
            row.get("fecha_movimiento"), row.get("fecha_renovacion"), row.get("estado"), row.get("estado_manual"),
            row.get("tipo_plla"), row.get("area"), row.get("cargo"), row.get("operacion"), row.get("estado_visual"),
            row.get("archivo_origen"), row.get("fecha_importacion"),
        ])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"reporte_epps_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.template_filter("estado_badge")
def estado_badge(value):
    value = (value or "").strip().upper()
    if value == "VENCIDO":
        return "badge-red"
    if value == "POR VENCER":
        return "badge-yellow"
    if value in {"ENTREGADO", "VIGENTE"}:
        return "badge-green"
    return "badge-gray"


@app.context_processor
def inject_now():
    return {"current_year": datetime.now().year}


init_db()


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
